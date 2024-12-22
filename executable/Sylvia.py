import pandas as pd
import random
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Funktionen

# Importieren aller Aufgaben aus einem Excel-File
def import_tasks(file_path):
    tasks_df = pd.read_excel(file_path)
    return tasks_df['Task'].tolist()

# Aktuellen Monat einlesen und Überschriften erstellen
# Titel "Tätigkeitsliste <eingegebener Monat>/<aktuelles Jahr>"
def get_month():
    month = int(input("Bitte geben Sie den Monat (Zahl 1-12) ein: "))
    year = datetime.now().year
    return f"Tätigkeitsliste für den Monat {month:02d}/{year}", month, year

# Berechnung der variablen Feiertage
def get_variable_holidays(year):
    easter = calculate_easter(year)
    holidays = {
        'Ostersonntag': easter,
        'Ostermontag': easter + timedelta(days=1),
        'Christi Himmelfahrt': easter + timedelta(days=39),
        'Pfingstmontag': easter + timedelta(days=50),
        'Fronleichnam': easter + timedelta(days=60)
    }
    return holidays

# Berechnung des Ostersonntags
def calculate_easter(year):
    # Computus algorithmus zur Berechnung von Ostern
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return datetime(year, month, day)

# Aktuelle Monatsübersicht inklusive der Feiertage in Bayern erstellen
def create_month(month, year):
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
    
    month_days = [start_date + timedelta(days=i) for i in range((end_date - start_date).days)]

    variable_holidays = get_variable_holidays(year)
    
    # Feiertage Bayern
    feiertage_bayern = {
        datetime(year, 1, 1): "Neujahr",
        datetime(year, 1, 6): "Heilige Drei Könige",
        variable_holidays['Ostersonntag']: "Ostersonntag",
        variable_holidays['Ostermontag']: "Ostermontag",
        datetime(year, 5, 1): "Tag der Arbeit",
        variable_holidays['Christi Himmelfahrt']: "Christi Himmelfahrt",
        variable_holidays['Pfingstmontag']: "Pfingstmontag",
        variable_holidays['Fronleichnam']: "Fronleichnam",
        datetime(year, 8, 15): "Mariä Himmelfahrt",
        datetime(year, 10, 3): "Tag der Deutschen Einheit",
        datetime(year, 11, 1): "Allerheiligen",
        datetime(year, 12, 25): "1. Weihnachtsfeiertag",
        datetime(year, 12, 26): "2. Weihnachtsfeiertag"
    }
    
    return month_days, feiertage_bayern

# Soll-Arbeitsstunden und tatsächlich geleistete Stunden eintragen
def working_hours(month_days, feiertage_bayern, urlaubstage):
    working_days = [day for day in month_days if day.weekday() < 5 and day not in feiertage_bayern and day not in urlaubstage]
    
    soll_hours = 8 * len(working_days)
    ist_hours = 8 * len(working_days)
    return soll_hours, ist_hours, working_days

# Urlaubstage einlesen
def get_urlaubstage(month, year):
    urlaub = input("Hattest du in dem Monat Urlaub? (j/n): ").strip().lower()
    print()
    urlaubstage = []
    if urlaub == 'j':
        tage = input("Bitte geben Sie die Urlaubstage (z.B. 9, 12, 15-17) ein: ")
        tage_list = tage.split(',')
        for tag in tage_list:
            if '-' in tag:
                start, end = tag.split('-')
                start, end = int(start.strip()), int(end.strip())
                urlaubstage.extend([datetime(year, month, day) for day in range(start, end + 1)])
            else:
                urlaubstage.append(datetime(year, month, int(tag.strip())))
    return urlaubstage

# Aufgaben zufällig auf alle Arbeitstage verteilen
def create_tasks(tasks, working_days):
    task_distribution = {}
    for day in working_days:
        num_tasks = random.randint(6, 7)
        task_distribution[day] = random.sample(tasks, num_tasks)
    return task_distribution

# Funktion zur Umwandlung des Wochentags in das Format Mo, Di, Mi etc.
def get_short_weekday(weekday):
    return ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"][weekday]

# Monatsliste auf dem Bildschirm ausdrucken
def print_tasks(header, month_days, task_distribution, feiertage_bayern, urlaubstage, soll_hours, ist_hours):
    lastname = "Prigge"
    firstname = "Sylvia"
    mitarbeiternummer = "XXXXXXX"
    wochenarbeitszeit = 40
    resturlaub = 7.5
    stand_date = datetime.now().strftime("%d.%m.%Y")

    # Fett gedruckter Header
    print('\033[1m' + header + '\033[0m')
    print('\033[1m' + '' + '\033[0m')  # Leere Zeile
    print("Nachname: " + lastname)
    print("Vorname: " + firstname)
    print("Mitarbeiternummer: " + mitarbeiternummer)
    print("Wochenarbeitszeit: {} h/Wo".format(wochenarbeitszeit))
    print("Resturlaub: {} Tage".format(resturlaub))
    print("Stand: {}".format(stand_date))
    print("\nDatum   Dauer    Wochentag    Tätigkeit1  Tätigkeit2  Tätigkeit3  Tätigkeit4  Tätigkeit5  Tätigkeit6  Tätigkeit7")
    print("="*120)

    for day in month_days:
        day_str = day.strftime("%d.%m.")
        weekday_str = get_short_weekday(day.weekday())
        tasks = task_distribution.get(day, [""] * 7)  # Fülle leere Aufgaben mit leeren Strings

        if day in feiertage_bayern:
            print(f"\033[91m{day_str}  0:00     {weekday_str:<2}  {feiertage_bayern[day]}\033[0m")
        elif day in urlaubstage:
            print(f"\033[91m{day_str}  0:00     {weekday_str:<2}  Urlaub\033[0m")
        elif day.weekday() >= 5:
            print(f"\033[94m{day_str}  0:00     {weekday_str:<2}  Wochenende\033[0m")
        else:
            print(f"{day_str}  8:00     {weekday_str:<2}  " + "  ".join(tasks))

    print("\nGesamt:")
    print("Sollarbeitszeit {:02d}:00 h".format(soll_hours))
    print("Ist-Arbeitszeit {:02d}:00 h".format(ist_hours))

    difference = ist_hours - soll_hours
    print("Differenz {:02d}:00 h".format(difference))

    print("\nDie Tätigkeitsliste {}/{} wurde erstellt und als excel file output.xlsx gespeichert.".format(month_days[0].month, month_days[0].year))

# Monatsliste als Excel exportieren
def export_tasks(header, month_days, task_distribution, feiertage_bayern, urlaubstage, soll_hours, ist_hours, output_path):
    data = []

    for day in month_days:
        day_str = day.strftime("%d.%m.")
        weekday_str = get_short_weekday(day.weekday())
        tasks = task_distribution.get(day, [""] * 7)  # Fülle leere Aufgaben mit leeren Strings

        if day in feiertage_bayern:
            task1 = feiertage_bayern[day]
            data.append([day_str, "0:00", weekday_str, "0:00", "Feiertag", task1] + [""] * 6)
        elif day in urlaubstage:
            task1 = "Urlaub"
            data.append([day_str, "0:00", weekday_str, "0:00", "Urlaub", task1] + [""] * 6)
        elif day.weekday() >= 5:
            task1 = "Wochenende"
            data.append([day_str, "0:00", weekday_str, "0:00", "Wochenende", task1] + [""] * 6)
        else:
            data.append([day_str, "8:00", weekday_str, "8:00", "Arbeitstag"] + tasks)

    columns = ['Datum', 'Dauer', 'Wochentag', 'Arbeitszeit', 'TagTyp', 'Tätigkeit1', 'Tätigkeit2', 'Tätigkeit3', 'Tätigkeit4', 'Tätigkeit5', 'Tätigkeit6', 'Tätigkeit7']
    df = pd.DataFrame(data, columns=columns)

    # Hinzufügen der leeren Zeilen und der Gesamtzeilen muss die gleiche Anzahl an Spalten haben
    empty_row = [''] * len(columns)
    total_row = ['Gesamt:', '', '', '', '']
    soll_row = ['Sollarbeitszeit:', '{:02d}:00'.format(soll_hours), '', '', '']
    ist_row = ['Ist-Arbeitszeit:', '{:02d}:00'.format(ist_hours), '', '', '']
    difference = ist_hours - soll_hours
    diff_row = ['Differenz:', '{:02d}:00'.format(difference), '', '', '']

    df.loc[len(df)] = empty_row  # Leere Zeile
    df.loc[len(df)] = total_row + [''] * (len(columns) - len(total_row))
    df.loc[len(df)] = soll_row + [''] * (len(columns) - len(soll_row))
    df.loc[len(df)] = ist_row + [''] * (len(columns) - len(ist_row))
    df.loc[len(df)] = diff_row + [''] * (len(columns) - len(diff_row))

    # Speichern der Excel-Datei mit Pandas (vorläufig)
    df.to_excel(output_path, index=False)

    # Laden der gespeicherten Datei mit openpyxl zur Bearbeitung
    wb = load_workbook(output_path)
    ws = wb.active

    # Formatierungen
    header_font = Font(bold=True, size=14)
    info_font = Font(size=11)
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    right_alignment = Alignment(horizontal="right")

    # Mitarbeiternamen und andere Informationen einfügen
    lastname = "Prigge"
    firstname = "Sylvia"
    mitarbeiternummer = "XXXXXXX"
    wochenarbeitszeit = 40
    resturlaub = 7.5
    stand_date = datetime.now().strftime("%d.%m.%Y")

    # Einfügen der Überschrift und zusätzlichen Informationen
    ws.insert_rows(1, amount=10)
    headers = [
        header,
        "",  # Leere Zeile
        f"Nachname: {lastname}",
        f"Vorname: {firstname}",
        f"Mitarbeiternummer: {mitarbeiternummer}",
        f"Wochenarbeitszeit: {wochenarbeitszeit} h/Wo",
        f"Resturlaub: {resturlaub} Tage",
        f"Stand: {stand_date}"
    ]

    # Überschrift fett und Schriftgröße 14
    cell = ws.cell(row=1, column=1, value=headers[0])
    cell.font = header_font

    # Restliche Informationen Schriftgröße 11
    for i, line in enumerate(headers[1:], start=2):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = info_font

    # Formatierung der Zeilen basierend auf dem TagTyp
    for row in ws.iter_rows(min_row=11, min_col=1, max_col=len(columns), max_row=ws.max_row):
        if row[4].value in ["Feiertag", "Urlaub"]:
            for cell in row:
                cell.fill = red_fill
        elif row[4].value == "Wochenende":
            for cell in row:
                cell.fill = blue_fill

        # Rahmen erstellen und Ausrichtung für Datum und Dauer setzen
        for cell in row:
            cell.border = thin_border
        row[0].alignment = right_alignment  # Datum
        row[1].alignment = right_alignment  # Dauer

    # Entfernen der Hilfsspalte mit dem TagTyp
    ws.delete_cols(5)

    # Speichern Sie die Änderungen
    wb.save(output_path)

# Hauptprogramm
def main():
    file_path = 'Projekte_SCL.xlsx'
    output_path = 'output.xlsx'

    tasks = import_tasks(file_path)
    header, month, year = get_month()
    urlaubstage = get_urlaubstage(month, year)
    month_days, feiertage_bayern = create_month(month, year)
    soll_hours, ist_hours, working_days = working_hours(month_days, feiertage_bayern, urlaubstage)
    task_distribution = create_tasks(tasks, working_days)

    print_tasks(header, month_days, task_distribution, feiertage_bayern, urlaubstage, soll_hours, ist_hours)
    export_tasks(header, month_days, task_distribution, feiertage_bayern, urlaubstage, soll_hours, ist_hours, output_path)

if __name__ == '__main__':
    main()
