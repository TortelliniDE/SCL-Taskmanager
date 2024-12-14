import pandas as pd
import random
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Funktionen

# Importieren aller Aufgaben aus einem Excel-File
def import_tasks(file_path):
    tasks_df = pd.read_excel(file_path)
    return tasks_df['Task'].tolist()

# Aktuellen Monat einlesen und Überschriften erstellen
# Titel "Tätigkeitsliste <eingegebener Monat>/<aktuelles Jahr>"
def get_month():
    month = int(input("Bitte geben Sie den Monat (Zahl 1-12) ein: "))
    year = datetime.now().year
    return f"\nTätigkeitsliste {month:02d}/{year}", month, year

# Berechnung der variablen Feiertage
def get_variable_holidays(year):
    easter = calculate_easter(year)
    holidays = {
        'Ostersonntag': easter,
        'Ostermontag': easter + timedelta(days=1),
        'Christi Himmelfahrt': easter + timedelta(days=39),
        'Pfingstmontag': easter + timedelta(days=50),
        'Fronleichnam': easter + timedelta(days=60)
    }
    return holidays

# Berechnung des Ostersonntags
def calculate_easter(year):
    # Computus algorithmus zur Berechnung von Ostern
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return datetime(year, month, day)

# Aktuelle Monatsübersicht inklusive der Feiertage in Bayern erstellen
def create_month(month, year):
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
    
    month_days = [start_date + timedelta(days=i) for i in range((end_date - start_date).days)]

    variable_holidays = get_variable_holidays(year)
    
    # Feiertage Bayern
    feiertage_bayern = {
        datetime(year, 1, 1): "Neujahr",
        datetime(year, 1, 6): "Heilige Drei Könige",
        variable_holidays['Ostersonntag']: "Ostersonntag",
        variable_holidays['Ostermontag']: "Ostermontag",
        datetime(year, 5, 1): "Tag der Arbeit",
        variable_holidays['Christi Himmelfahrt']: "Christi Himmelfahrt",
        variable_holidays['Pfingstmontag']: "Pfingstmontag",
        variable_holidays['Fronleichnam']: "Fronleichnam",
        datetime(year, 8, 15): "Mariä Himmelfahrt",
        datetime(year, 10, 3): "Tag der Deutschen Einheit",
        datetime(year, 11, 1): "Allerheiligen",
        datetime(year, 12, 25): "1. Weihnachtsfeiertag",
        datetime(year, 12, 26): "2. Weihnachtsfeiertag"
    }
    
    return month_days, feiertage_bayern

# Soll-Arbeitsstunden und tatsächlich geleistete Stunden eintragen
def working_hours(month_days, feiertage_bayern):
    working_days = [day for day in month_days if day.weekday() < 5 and day not in feiertage_bayern]
    
    soll_hours = 8 * len(working_days)
    ist_hours = 8 * len(working_days)
    return soll_hours, ist_hours, working_days

# Aufgaben zufällig auf alle Arbeitstage verteilen
def create_tasks(tasks, working_days):
    task_distribution = {}
    for day in working_days:
        num_tasks = random.randint(3, 5)
        task_distribution[day] = random.sample(tasks, num_tasks)
    return task_distribution

# Monatsliste auf dem Bildschirm ausdrucken
def print_tasks(header, month_days, task_distribution, feiertage_bayern, soll_hours, ist_hours):
    mitarbeiternummer = "SCL-4711"
    wochenarbeitszeit = 40
    resturlaub = 7.5
    stand_date = datetime.now().strftime("%d.%m.%Y")

    # Fett gedruckter Header
    print('\033[1m' + header + '\033[0m')
    print('\033[1m' + '' + '\033[0m')  # Leere Zeile
    print("\nMitarbeiternummer: " + mitarbeiternummer)
    print("Wochenarbeitszeit: {} h/Wo".format(wochenarbeitszeit))
    print("Resturlaub: {} Tage".format(resturlaub))
    print("Stand: {}".format(stand_date))
    print("\nDatum   Dauer    Wochentag    Tätigkeiten")
    print("="*120)

    for day in month_days:
        day_str = day.strftime("%d.%m.")
        weekday_str = day.strftime("%A")
        if day in feiertage_bayern:
            print(f"\033[94m{day_str}  0:00     {weekday_str:<10}  {feiertage_bayern[day]}\033[0m")
        elif day.weekday() >= 5:
            print(f"\033[94m{day_str}  0:00     {weekday_str:<10}  Wochenende\033[0m")
        elif day in task_distribution:
            tasks = ", ".join(task_distribution[day])
            print(f"{day_str}  8:00     {weekday_str:<10}  {tasks}")
        else:
            print(f"{day_str}  0:00     {weekday_str:<10}  Wochenende")

    print("\nGesamt")
    print("Sollarbeitszeit {:02d}:00 h".format(soll_hours))
    print("Ist-Arbeitszeit {:02d}:00 h".format(ist_hours))

    difference = ist_hours - soll_hours
    print("Differenz {:02d}:00 h".format(difference))

    print("\nDie Tätigkeitsliste {}/{} wurde erstellt.".format(month_days[0].month, month_days[0].year))

# Monatsliste als Excel exportieren
def export_tasks(header, month_days, task_distribution, feiertage_bayern, soll_hours, ist_hours, output_path):
    data = []

    for day in month_days:
        day_str = day.strftime("%d.%m.")
        weekday_str = day.strftime("%A")
        if day in feiertage_bayern:
            data.append((day_str, "0:00", weekday_str, feiertage_bayern[day], "Feiertag"))
        elif day.weekday() >= 5:
            data.append((day_str, "0:00", weekday_str, "Wochenende", "Wochenende"))
        else:
            tasks = ", ".join(task_distribution.get(day, []))
            data.append((day_str, "8:00", weekday_str, tasks, "Arbeitstag"))

    df = pd.DataFrame(data, columns=['Datum', 'Dauer', 'Wochentag', 'Tätigkeiten', 'TagTyp'])
    df.loc[len(df)] = ['', '', '', '', '']  # Leere Zeile
    df.loc[len(df)] = ['Gesamt', '', '', '', '']
    df.loc[len(df)] = ['Sollarbeitszeit', '{:02d}:00'.format(soll_hours), '', '', '']
    df.loc[len(df)] = ['Ist-Arbeitszeit', '{:02d}:00'.format(ist_hours), '', '', '']
    
    difference = ist_hours - soll_hours
    df.loc[len(df)] = ['Differenz', '{:02d}:00'.format(difference), '', '', '']

    # Speichern der Excel-Datei mit Pandas (vorläufig)
    df.to_excel(output_path, index=False)

    # Laden der gespeicherten Datei mit openpyxl zur Bearbeitung
    wb = load_workbook(output_path)
    ws = wb.active

    # Fett gedruckter Header
    header_font = Font(bold=True)
    # Blau gefärbte Zellen
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Einfügen der Überschrift und zusätzlichen Informationen
    ws.insert_rows(1, amount=8)
    headers = [
        header,
        "",  # Leere Zeile
        "Mitarbeiternummer: SCL-4711",
        "Wochenarbeitszeit: 40 h/Wo",
        "Resturlaub: 7.5 Tage",
        "Stand: {}".format(datetime.now().strftime("%d.%m.%Y"))
    ]

    for i, line in enumerate(headers, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = header_font

    # Formatierung der Zeilen basierend auf dem TagTyp
    for row in ws.iter_rows(min_row=9, min_col=1, max_col=5, max_row=ws.max_row):
        if row[4].value in ["Feiertag", "Wochenende"]:
            for cell in row:
                cell.fill = blue_fill

    # Entfernen der Hilfsspalte mit dem TagTyp
    ws.delete_cols(5)

    # Speichern Sie die Änderungen
    wb.save(output_path)

# Hauptprogramm
def main():
    file_path = 'Projekte_SCL.xlsx'
    output_path = 'output.xlsx'

    tasks = import_tasks(file_path)
    header, month, year = get_month()
    month_days, feiertage_bayern = create_month(month, year)
    soll_hours, ist_hours, working_days = working_hours(month_days, feiertage_bayern)
    task_distribution = create_tasks(tasks, working_days)

    print_tasks(header, month_days, task_distribution, feiertage_bayern, soll_hours, ist_hours)
    export_tasks(header, month_days, task_distribution, feiertage_bayern, soll_hours, ist_hours, output_path)

if __name__ == '__main__':
    main()
